from bs4 import BeautifulSoup
from selenium import webdriver
import openpyxl

'''
1. Get all url of products
2. Get product data from each url 
3. Write to excel
'''

driver = webdriver.Chrome()
url = "https://komehyo.jp/chanel/"
driver.get(url)
html = driver.page_source.encode('utf-8')
soup = BeautifulSoup(html, "lxml")
item_name_list = ["image", "title", "category", "Item_Id_in_komehyo", "Sales_price", "Model_No", "Material", "Color", "Size", "Stock_store", "Gender_type", "product_page_url"]
url_links = []
output = []

# 1. Get all url links
def get_all_url_links():
    page_num = 1
    while(can_go_to_next_page(page_num)):
        get_url_links()
        page_num += 1
        if page_num >= 2:
            break
    print("num of url links: ", len(url_links))
    return url_links

def can_go_to_next_page(page_num=1):
    soup = set_soup("{}?page={}".format(url, str(page_num)))
    num_of_products = soup.find("span", class_="p-txt p-txt--result-num").text

    if(num_of_products == "0"):
        return False
    return True

def get_url_links():
    products = soup.find_all("a",class_="p-link p-link--card")
    for product in products:
        url_links.append(product.get("href"))
    return url_links

def set_soup(url):
    try:
        driver.get(url)
        html = driver.page_source.encode('utf-8')
        soup = BeautifulSoup(html, "lxml")
        return soup
    except:
        print("Error: ", url)
        return None

# 2. Get product data from each url
def get_all_product_info():
    url_links = get_all_url_links()
    for url_link in url_links:
        page_url = "https://komehyo.jp" + url_link
        if set_soup(page_url) is not None:
            output.append(get_product_info(page_url))

    driver.close()
    driver.quit()
    print("num of products: ", len(output))
    return output

def get_product_info(product_page_url):
    soup = set_soup(product_page_url)
    non_table_contents = find_non_table_data(soup)
    table_contents = find_table_data(soup)
    item_list = non_table_contents + table_contents + [product_page_url]
    return item_list

def find_non_table_data(soup):
    title = find_non_table_data_helper(soup, "div", "p-product-name")
    Sales_price = find_non_table_data_helper(soup, "div", "p-block--selling-price")
    Item_Id_in_komehyo = find_non_table_data_helper(soup, "div", "p-product-code")
    Item_Id_in_komehyo = Item_Id_in_komehyo.replace("商品コード：", "")
    image = find_image(soup)
    category = find_category(soup)
    return [image, title, category, Item_Id_in_komehyo, Sales_price]

def find_non_table_data_helper(soup, tag, class_name):
    if soup.find(tag, class_=class_name) is not None:
        return soup.find(tag, class_=class_name).text.replace("\n", "")
    return ""

def find_category(soup):
    if soup.find_all("a", class_="c-breadcrumbs__item__link") is not None:
        category_list = soup.find_all("a", class_="c-breadcrumbs__item__link")
        category_str = ""
        for cat in category_list:
            category_str += cat.text.replace("\n", "") + " > "
        return category_str
    return ""

def find_image(soup):
    if soup.find("a", class_="js-picturemodal js-picturemodal1") is not None:
        return soup.find("a", class_="js-picturemodal js-picturemodal1").find("img").get("src")

def find_table_data(soup):
    if soup.find(class_="p-table__content") is not None:
        contents = soup.find(class_="p-table__content")
        Model_No = find_table_data_helper(contents, ["品番", "品番型式", "型式"])
        Material = find_table_data_helper(contents, ["素材"])
        Color = find_table_data_helper(contents, ["カラー"])
        Gender_type = find_table_data_helper(contents, ["性別タイプ"])
        Stock_store = find_table_data_helper(contents, ["在庫店舗"])
        Size = contents.find("a", class_="p-link js-modal p-link--help").find("span", text="サイズ").parent.parent.find_next_sibling("td").text.replace("\n", "")
        return [Model_No, Material, Color, Size, Stock_store, Gender_type]
    return ["", "", "", "", "", "", ""]

def find_table_data_helper(contents, table_head):
    for head in table_head:
        if contents.find("th", text=head) is not None:
            row_description = contents.find("th", text=head).find_next_sibling("td").text.replace("\n", "")
            return row_description
    return ""

# 3. Write to excel
def write_to_excel():
    wb = openpyxl.Workbook()
    wb.save('komehyo_chanel.xlsx')
    sheet = wb['Sheet']

    for column, item_name in enumerate(item_name_list):
        sheet.cell(row=1, column=column+1).value = item_name

    for row in range(len(output)):
        for col in range(len(output[row])):
            sheet.cell(row=row+2, column=col+1).value = output[row][col]
    wb.save('komehyo_chanel.xlsx')

# Execute functions
get_all_product_info()
write_to_excel()